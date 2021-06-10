import tkinter 
import os
from tkinter.ttk import Combobox
from tkinter import filedialog
from tkinter import *	
from tkinter import messagebox		
#from openpyxl import workbook
from openpyxl import load_workbook
import sys

#Crear ventana modelo SWMM

def ventana_modelo():

	global ventmodel
	ventmodel=tkinter.Toplevel()
	ventmodel.geometry("800x600+500+250")
	ventmodel.resizable(0,0)
	ventmodel.title("SWMM-EXCEL Alpha - CREADOR DE MODELO ALCANTARILLADO EPASWMM")
	ventmodel.iconbitmap(r"C:\Users\Ing Juan Manuel\Documents\Sublime Text Ejercicios\SWMM-EXCEL\Icono.ico")

	#creador botones

	botoncargararchivo = tkinter.Button(ventmodel, text = "1.Cargar Archivo", command = cargararchivo).place(x=10,y=20)
	botoncargarlistasenorden = tkinter.Button(ventmodel, text = "2.Cargar listas en orden", command = cargarlistasenorden).place(x=120,y=20)
	botoncargarlistas = tkinter.Button(ventmodel, text = "3.Cargar listas", command = cargarlistas).place(x=270,y=20)
	botoncrearmodelo = tkinter.Button(ventmodel, text = "4.Crear Modelo", command = crearmodelo).place(x=370,y=20)
	botoncerrar = tkinter.Button(ventmodel, text = "5.Cerrar", command = ventmodel.destroy).place(x=470,y=20)


	#Crear Barra Menu

	barraMenumodel=Menu(ventmodel)

	#Crear los menus

	mnucrearmodelo = Menu(barraMenumodel, tearoff = 0)
	mnuarchivo = Menu(barraMenumodel, tearoff = 0)
	mnuherramientas=Menu(barraMenumodel, tearoff = 0)

	#Crear los comandos de los menus

	mnucrearmodelo.add_command(label="Crear modelo SWMM-EXCEL",command=ventana_modelo)

	#####

	mnuarchivo.add_command(label="Cargar Archivo Excel",command=cargararchivo)
	mnuarchivo.add_separator()
	mnuarchivo.add_command(label="Salir",command=ventmodel.destroy)

	######


	mnuherramientas.add_command(label="Cargar listas en orden",command=cargarlistasenorden)
	mnuherramientas.add_command(label="Cargar listas por defecto",command=cargarlistas)
	mnuherramientas.add_separator()
	mnuherramientas.add_command(label="Crear modelo",command=crearmodelo)

	#Agregar los menus a la barra de menus

	barraMenumodel.add_cascade(label="Archivo",menu=mnuarchivo)
	barraMenumodel.add_cascade(label="Herramientas",menu=mnuherramientas)
	
	#Indicamos que la barra de menus estará en la ventana

	ventmodel.config(menu=barraMenumodel)

	
	#creador etiquetas

	autor = tkinter.Label(ventmodel, text= "Ing. Juan Manuel Pérez - Bogotá Colombia - Youtube: Juan Manuel Perez Hidráulica").place(x=200,y=550)
	etiquetahojapozos = tkinter.Label(ventmodel, text= "Hoja Pozos").place(x=10,y=70)
	IDpozo = tkinter.Label(ventmodel, text= "ID Pozo").place(x=10,y=100)
	Estepozo = tkinter.Label(ventmodel, text= "Coord. Este Pozo (m)").place(x=10,y=130)
	Nortepozo = tkinter.Label(ventmodel, text= "Coord. Norte Pozo (m)").place(x=10,y=160)
	Cotaterrenopozo = tkinter.Label(ventmodel, text= "Cota Terreno Pozo").place(x=10,y=190)
	Cotafondopozo = tkinter.Label(ventmodel, text= "Cota Fondo Pozo").place(x=10,y=220)
	etiquetahojatramos = tkinter.Label(ventmodel, text= "Hoja Tramos").place(x=10,y=250)
	pozoinicial = tkinter.Label(ventmodel, text= "Pozo Inicial").place(x=10,y=280)
	pozofinal = tkinter.Label(ventmodel, text= "Pozo Final").place(x=10,y=310)
	diametro = tkinter.Label(ventmodel, text= "Diametro tramo (m)").place(x=10,y=340)
	rugosidad = tkinter.Label(ventmodel, text= "Manning").place(x=10,y=370)
	claveinicial = tkinter.Label(ventmodel, text= "Cota Clave Inicial").place(x=10,y=400)
	clavefinal = tkinter.Label(ventmodel, text= "Cota Clave Final").place(x=10,y=430)
	Caudal = tkinter.Label(ventmodel, text= "Caudal (L/s)").place(x=10,y=460)
	etiquetahojadescarga = tkinter.Label(ventmodel, text= "Hoja Descargas").place(x=400,y=70)
	IDdescarga = tkinter.Label(ventmodel, text= "ID Descarga").place(x=400,y=100)
	Estedescarga = tkinter.Label(ventmodel, text= "Coord. Este Descarga (m)").place(x=400,y=130)
	Nortedescarga = tkinter.Label(ventmodel, text= "Coord. Norte Descarga (m)").place(x=400,y=160)
	cotafondodescarga = tkinter.Label(ventmodel, text= "Cota Fondo Descarga").place(x=400,y=190)

#Si se ingresa el caudal acumulado con el siguiente comando se calcula el caudal puntual a ingresar al modelo

def calcular_caudal_puntual():

	for row in listahojatramos.iter_rows(min_col=columnacaudal, max_col=columnacaudal):
		caudalcargado.extend([celda.value for celda in row])

	del(caudalcargado[0])

	global caudalcargadopuntual
	caudalcargadopuntual=[]

	for i in range(longitud_lista_tramos):

		caudal_puntual=caudalcargado[i]

		for j in range(longitud_lista_tramos):

			if (pozoinicialcargado[i]==pozofinalcargado[j]):

				caudal_puntual=caudal_puntual-caudalcargado[j]

				#print("caudal puntual "+str(caudal_puntual))
		
		caudalcargadopuntual.extend([caudal_puntual])

		
#cargar archivo

def cargararchivo():


	messagebox.showinfo(parent=ventmodel, message="Tenga en cuenta las siguientes recomendaciones para cargar el archivo:\n1.El archivo a cargar debe ser de tipo excel (xlsx)\n2.El archivo de excel debe terner mínimo 3 hojas (tramos,pozos,descargas)\n3.No deben existir filas en blanco entre los datos del archivo de excel\n4.El archivo debe contener la totalidad de la información requerida por el programa\n5.Las columnas deben contar con titulos en los encabezados\n6.Para más información visita https://www.youtube.com/c/JUANPEREZ", title="Alerta")
	global hojaexcel
	global hojapozos
	archivo = filedialog.askopenfilename(parent=ventmodel,title="Abrir")
	archivo2 = os.path.abspath(archivo)
	hojaexcel = load_workbook(archivo2)
	#print(hojaexcel)
	hojapozos = hojaexcel.get_sheet_names()
	#print(hojapozos)

	#crear lista de hojas de excel pozos

	global listahojasexcel
	listahojasexcel = Combobox(ventmodel)
	listahojasexcel.place(x=160,y=70)
	listahojasexcel["values"] = tuple(hojapozos)
	listahojasexcel.current(0)

	#crear lista de hojas de excel tramos

	global listahojasexcel2
	listahojasexcel2 = Combobox(ventmodel)
	listahojasexcel2.place(x=160,y=250)
	listahojasexcel2["values"] = tuple(hojapozos)
	listahojasexcel2.current(1)

	#crear lista de hojas de excel descargas

	global listahojasexcel3
	listahojasexcel3 = Combobox(ventmodel)
	listahojasexcel3.place(x=560,y=70)
	listahojasexcel3["values"] = tuple(hojapozos)
	listahojasexcel3.current(2)

def cargarlistas():

	#crear lista de titulos de hoja de pozos

	global hojapozosseleccionada #global crea variables globales 
	hojapozosseleccionada = listahojasexcel.get()
	global listahojapozos
	listahojapozos = hojaexcel[hojapozosseleccionada]
	global listatitulospozos
	listatitulospozos = []

	for columna in listahojapozos.iter_rows(min_row=1, max_row=1):
		listatitulospozos.extend([celda.value for celda in columna])

	#crear lista de titulos de hoja de tramos

	global hojatramosseleccionada #global crea variables globales 
	hojatramosseleccionada = listahojasexcel2.get()
	global listahojatramos
	listahojatramos = hojaexcel[hojatramosseleccionada]
	global listatitulostramos
	listatitulostramos = []

	for columna in listahojatramos.iter_rows(min_row=1, max_row=1):
		listatitulostramos.extend([celda.value for celda in columna])



	#crear lista de titulos de hoja de descargas

	global hojadescargaseleccionada #global crea variables globales 
	hojadescargaseleccionada = listahojasexcel3.get()
	global listahojadescarga
	listahojadescarga = hojaexcel[hojadescargaseleccionada]
	global listatitulosdescarga
	listatitulosdescarga = []

	for columna in listahojadescarga.iter_rows(min_row=1, max_row=1):
		listatitulosdescarga.extend([celda.value for celda in columna])

	#crear listas titulos Pozos

	global listaIDpozo
	listaIDpozo = Combobox(ventmodel)
	listaIDpozo.place(x=160,y=100)
	listaIDpozo["values"] = tuple(listatitulospozos)
	listaIDpozo.current(0)

	global listaEstepozo
	listaEstepozo = Combobox(ventmodel)
	listaEstepozo.place(x=160,y=130)
	listaEstepozo["values"] = tuple(listatitulospozos)
	listaEstepozo.current(0)

	global listaNortepozo
	listaNortepozo = Combobox(ventmodel)
	listaNortepozo.place(x=160,y=160)
	listaNortepozo["values"] = tuple(listatitulospozos)
	listaNortepozo.current(0)

	global listaCotaterrenopozo
	listaCotaterrenopozo = Combobox(ventmodel)
	listaCotaterrenopozo.place(x=160,y=190)
	listaCotaterrenopozo["values"] = tuple(listatitulospozos)
	listaCotaterrenopozo.current(0)

	global listaCotafondopozo
	listaCotafondopozo = Combobox(ventmodel)
	listaCotafondopozo.place(x=160,y=220)
	listaCotafondopozo["values"] = tuple(listatitulospozos)
	listaCotafondopozo.current(0)

	#crear listas titulos tramos

	global listapozoinicial
	listapozoinicial = Combobox(ventmodel)
	listapozoinicial.place(x=160,y=280)
	listapozoinicial["values"] = tuple(listatitulostramos)
	listapozoinicial.current(0)

	global listapozofinal
	listapozofinal = Combobox(ventmodel)
	listapozofinal.place(x=160,y=310)
	listapozofinal["values"] = tuple(listatitulostramos)
	listapozofinal.current(0)

	global listadiametro
	listadiametro = Combobox(ventmodel)
	listadiametro.place(x=160,y=340)
	listadiametro["values"] = tuple(listatitulostramos)
	listadiametro.current(0)

	global listarugosidad
	listarugosidad = Combobox(ventmodel)
	listarugosidad.place(x=160,y=370)
	listarugosidad["values"] = tuple(listatitulostramos)
	listarugosidad.current(0)

	global listaclaveinicial
	listaclaveinicial = Combobox(ventmodel)
	listaclaveinicial.place(x=160,y=400)
	listaclaveinicial["values"] = tuple(listatitulostramos)
	listaclaveinicial.current(0)

	global listaclavefinal
	listaclavefinal = Combobox(ventmodel)
	listaclavefinal.place(x=160,y=430)
	listaclavefinal["values"] = tuple(listatitulostramos)
	listaclavefinal.current(0)

	global listacaudal
	listacaudal = Combobox(ventmodel)
	listacaudal.place(x=160,y=460)
	listacaudal["values"] = tuple(listatitulostramos)
	listacaudal.current(0)

	#Crea radio button para caudal
	global selec
	selec=IntVar()
	caudal_puntual=Radiobutton(ventmodel,text="Caudal Puntual",value=1,variable=selec).place(x=350,y=460)
	caudal_acumulado=Radiobutton(ventmodel,text="Caudal Acumulado",value=2,variable=selec).place(x=350,y=490)

	#crear listas titulos descargas

	global listaIDdescarga
	listaIDdescarga = Combobox(ventmodel)
	listaIDdescarga.place(x=560,y=100)
	listaIDdescarga["values"] = tuple(listatitulosdescarga)
	listaIDdescarga.current(0)

	global listaEstedescarga
	listaEstedescarga = Combobox(ventmodel)
	listaEstedescarga.place(x=560,y=130)
	listaEstedescarga["values"] = tuple(listatitulosdescarga)
	listaEstedescarga.current(0)

	global listaNortedescarga
	listaNortedescarga = Combobox(ventmodel)
	listaNortedescarga.place(x=560,y=160)
	listaNortedescarga["values"] = tuple(listatitulosdescarga)
	listaNortedescarga.current(0)

	global listacotafondodescarga
	listacotafondodescarga = Combobox(ventmodel)
	listacotafondodescarga.place(x=560,y=190)
	listacotafondodescarga["values"] = tuple(listatitulosdescarga)
	listacotafondodescarga.current(0)

def cargarlistasenorden():

	#crear lista de titulos de hoja de pozos

	global hojapozosseleccionada #global crea variables globales 
	hojapozosseleccionada = listahojasexcel.get()
	global listahojapozos
	listahojapozos = hojaexcel[hojapozosseleccionada]
	global listatitulospozos
	listatitulospozos = []

	for columna in listahojapozos.iter_rows(min_row=1, max_row=1):
		listatitulospozos.extend([celda.value for celda in columna])

	#crear lista de titulos de hoja de tramos

	global hojatramosseleccionada #global crea variables globales 
	hojatramosseleccionada = listahojasexcel2.get()
	global listahojatramos
	listahojatramos = hojaexcel[hojatramosseleccionada]
	global listatitulostramos
	listatitulostramos = []

	for columna in listahojatramos.iter_rows(min_row=1, max_row=1):
		listatitulostramos.extend([celda.value for celda in columna])

	

	#crear lista de titulos de hoja de descargas

	global hojadescargaseleccionada #global crea variables globales 
	hojadescargaseleccionada = listahojasexcel3.get()
	global listahojadescarga
	listahojadescarga = hojaexcel[hojadescargaseleccionada]
	global listatitulosdescarga
	listatitulosdescarga = []

	for columna in listahojadescarga.iter_rows(min_row=1, max_row=1):
		listatitulosdescarga.extend([celda.value for celda in columna])
	
	#crear listas titulos Pozos

	global listaIDpozo
	listaIDpozo = Combobox(ventmodel)
	listaIDpozo.place(x=160,y=100)
	listaIDpozo["values"] = tuple(listatitulospozos)
	listaIDpozo.current(0)

	global listaEstepozo
	listaEstepozo = Combobox(ventmodel)
	listaEstepozo.place(x=160,y=130)
	listaEstepozo["values"] = tuple(listatitulospozos)
	listaEstepozo.current(1)

	global listaNortepozo
	listaNortepozo = Combobox(ventmodel)
	listaNortepozo.place(x=160,y=160)
	listaNortepozo["values"] = tuple(listatitulospozos)
	listaNortepozo.current(2)

	global listaCotaterrenopozo
	listaCotaterrenopozo = Combobox(ventmodel)
	listaCotaterrenopozo.place(x=160,y=190)
	listaCotaterrenopozo["values"] = tuple(listatitulospozos)
	listaCotaterrenopozo.current(3)

	global listaCotafondopozo
	listaCotafondopozo = Combobox(ventmodel)
	listaCotafondopozo.place(x=160,y=220)
	listaCotafondopozo["values"] = tuple(listatitulospozos)
	listaCotafondopozo.current(4)

	#crear listas titulos tramos

	global listapozoinicial
	listapozoinicial = Combobox(ventmodel)
	listapozoinicial.place(x=160,y=280)
	listapozoinicial["values"] = tuple(listatitulostramos)
	listapozoinicial.current(0)

	global listapozofinal
	listapozofinal = Combobox(ventmodel)
	listapozofinal.place(x=160,y=310)
	listapozofinal["values"] = tuple(listatitulostramos)
	listapozofinal.current(1)

	global listadiametro
	listadiametro = Combobox(ventmodel)
	listadiametro.place(x=160,y=340)
	listadiametro["values"] = tuple(listatitulostramos)
	listadiametro.current(2)

	global listarugosidad
	listarugosidad = Combobox(ventmodel)
	listarugosidad.place(x=160,y=370)
	listarugosidad["values"] = tuple(listatitulostramos)
	listarugosidad.current(3)

	global listaclaveinicial
	listaclaveinicial = Combobox(ventmodel)
	listaclaveinicial.place(x=160,y=400)
	listaclaveinicial["values"] = tuple(listatitulostramos)
	listaclaveinicial.current(4)

	global listaclavefinal
	listaclavefinal = Combobox(ventmodel)
	listaclavefinal.place(x=160,y=430)
	listaclavefinal["values"] = tuple(listatitulostramos)
	listaclavefinal.current(5)

	global listacaudal
	listacaudal = Combobox(ventmodel)
	listacaudal.place(x=160,y=460)
	listacaudal["values"] = tuple(listatitulostramos)
	listacaudal.current(6)

	#Crea radio button para caudal
	global selec
	selec=IntVar(ventmodel)
	caudal_puntual=Radiobutton(ventmodel,text="Caudal Puntual",value=1,variable=selec).place(x=350,y=460)
	caudal_acumulado=Radiobutton(ventmodel,text="Caudal Acumulado",value=2,variable=selec).place(x=350,y=490)

	#crear listas titulos descargas

	global listaIDdescarga
	listaIDdescarga = Combobox(ventmodel)
	listaIDdescarga.place(x=560,y=100)
	listaIDdescarga["values"] = tuple(listatitulosdescarga)
	listaIDdescarga.current(0)

	global listaEstedescarga
	listaEstedescarga = Combobox(ventmodel)
	listaEstedescarga.place(x=560,y=130)
	listaEstedescarga["values"] = tuple(listatitulosdescarga)
	listaEstedescarga.current(1)

	global listaNortedescarga
	listaNortedescarga = Combobox(ventmodel)
	listaNortedescarga.place(x=560,y=160)
	listaNortedescarga["values"] = tuple(listatitulosdescarga)
	listaNortedescarga.current(2)

	global listacotafondodescarga
	listacotafondodescarga = Combobox(ventmodel)
	listacotafondodescarga.place(x=560,y=190)
	listacotafondodescarga["values"] = tuple(listatitulosdescarga)
	listacotafondodescarga.current(3)

#crear modelo

def crearmodelo():

	#identificar el numero de la columna de acuerdo a lo seleccionado en las listas por el usuario
	
	tipo_caudal_seleccionado=selec.get()

	if (tipo_caudal_seleccionado==0):
		messagebox.showwarning(parent=ventmodel, message="Debe seleccionar primero el tipo de caudal a cargar (puntual o acumulado)", title="Alerta")

	else:
		nombrecolumnaidseleccionada = listaIDpozo.get()
		nombrecolumnaestepozoseleccionada = listaEstepozo.get()
		nombrecolumnanortepozoseleccionada = listaNortepozo.get()
		nombrecolumnacotaterrenoseleccionada = listaCotaterrenopozo.get()
		nombrecolumnacotafondoseleccionada = listaCotafondopozo.get()

		nombrecolumnapozoinicialseleccionada = listapozoinicial.get()
		nombrecolumnapozofinalseleccionada = listapozofinal.get()
		nombrecolumnadiametroseleccionada = listadiametro.get()
		nombrecolumnarugosidadseleccionada = listarugosidad.get()
		nombrecolumnaclaveinicialseleccionada = listaclaveinicial.get()
		nombrecolumnaclavefinalseleccionada = listaclavefinal.get()
		nombrecolumnacaudalseleccionada = listacaudal.get()

		nombrecolumnaiddescargaseleccionada = listaIDdescarga.get()
		nombrecolumnaestedescargaseleccionada = listaEstedescarga.get()	
		nombrecolumnanortedescargaseleccionada = listaNortedescarga.get()
		nombrecolumnacotafondodescargaseleccionada = listacotafondodescarga.get()

	#identifica el numero del campo de la lista donde se almacena el valor seleccionado por el usuario

		columnaid = listatitulospozos.index(nombrecolumnaidseleccionada)+1
		columnaestepozo = listatitulospozos.index(nombrecolumnaestepozoseleccionada)+1
		columnanortepozo = listatitulospozos.index(nombrecolumnanortepozoseleccionada)+1
		columnacotaterreno = listatitulospozos.index(nombrecolumnacotaterrenoseleccionada)+1
		columnacotafondo = listatitulospozos.index(nombrecolumnacotafondoseleccionada)+1

		columnapozoinicial = listatitulostramos.index(nombrecolumnapozoinicialseleccionada)+1
		columnapozofinal = listatitulostramos.index(nombrecolumnapozofinalseleccionada)+1
		columnadiametro = listatitulostramos.index(nombrecolumnadiametroseleccionada)+1
		columnarugosidad = listatitulostramos.index(nombrecolumnarugosidadseleccionada)+1
		columnaclaveinicial = listatitulostramos.index(nombrecolumnaclaveinicialseleccionada)+1
		columnaclavefinal = listatitulostramos.index(nombrecolumnaclavefinalseleccionada)+1
		global columnacaudal
		columnacaudal = listatitulostramos.index(nombrecolumnacaudalseleccionada)+1

		columnaiddescarga = listatitulosdescarga.index(nombrecolumnaiddescargaseleccionada)+1
		columnaestedescarga = listatitulosdescarga.index(nombrecolumnaestedescargaseleccionada)+1
		columnanortedescarga = listatitulosdescarga.index(nombrecolumnanortedescargaseleccionada)+1
		columnacotafondodescarga = listatitulosdescarga.index(nombrecolumnacotafondodescargaseleccionada)+1


		#print(columnaid)
		#print(columnaestepozo)
		#print(columnanortepozo)
		#print(columnacotaterreno)
		#print(columnacotafondo)

	#crear listas de valores 

		IDPozocargado = []
		estepozocargado = []
		nortepozocargado = []
		cotaterrenocargado = []
		cotafondocargado = []

		global pozoinicialcargado
		pozoinicialcargado = []
		global pozofinalcargado
		pozofinalcargado = []
		diametrocargado = []
		rugosidadcargado = []
		claveinicialcargado = []
		clavefinalcargado = []
		global caudalcargado
		caudalcargado = []

		IDdescargacargado = []
		estedescargacargado = []
		nortedescargacargado = []
		cotafondodescargacargado = []

	#cargar información a las listas por fila con base en la información de excel para pozos

		for row in listahojapozos.iter_rows(min_col=columnaid, max_col=columnaid):
			IDPozocargado.extend([celda.value for celda in row])

		for row in listahojapozos.iter_rows(min_col=columnaestepozo, max_col=columnaestepozo):
			estepozocargado.extend([celda.value for celda in row])

		for row in listahojapozos.iter_rows(min_col=columnanortepozo, max_col=columnanortepozo):
			nortepozocargado.extend([celda.value for celda in row])

		for row in listahojapozos.iter_rows(min_col=columnacotaterreno, max_col=columnacotaterreno):
			cotaterrenocargado.extend([celda.value for celda in row])

		for row in listahojapozos.iter_rows(min_col=columnacotafondo, max_col=columnacotafondo):
			cotafondocargado.extend([celda.value for celda in row])

	#cargar información a las listas por fila con base en la información de excel para tramos

		for row in listahojatramos.iter_rows(min_col=columnapozoinicial, max_col=columnapozoinicial):
			pozoinicialcargado.extend([celda.value for celda in row])	

		for row in listahojatramos.iter_rows(min_col=columnapozofinal, max_col=columnapozofinal):
			pozofinalcargado.extend([celda.value for celda in row])	

		for row in listahojatramos.iter_rows(min_col=columnadiametro, max_col=columnadiametro):
			diametrocargado.extend([celda.value for celda in row])	

		for row in listahojatramos.iter_rows(min_col=columnarugosidad, max_col=columnarugosidad):
			rugosidadcargado.extend([celda.value for celda in row])	

		for row in listahojatramos.iter_rows(min_col=columnaclaveinicial, max_col=columnaclaveinicial):
			claveinicialcargado.extend([celda.value for celda in row])	

		for row in listahojatramos.iter_rows(min_col=columnaclavefinal, max_col=columnaclavefinal):
			clavefinalcargado.extend([celda.value for celda in row])


	#cargar información a las listas por fila con base en la información de excel para descargas

		for row in listahojadescarga.iter_rows(min_col=columnaiddescarga, max_col=columnaiddescarga):
			IDdescargacargado.extend([celda.value for celda in row])	

		for row in listahojadescarga.iter_rows(min_col=columnaestedescarga, max_col=columnaestedescarga):
			estedescargacargado.extend([celda.value for celda in row])	

		for row in listahojadescarga.iter_rows(min_col=columnanortedescarga, max_col=columnanortedescarga):
			nortedescargacargado.extend([celda.value for celda in row])	

		for row in listahojadescarga.iter_rows(min_col=columnacotafondodescarga, max_col=columnacotafondodescarga):
			cotafondodescargacargado.extend([celda.value for celda in row])	

	#elimina el título de las listas

		del(IDPozocargado[0])
		del(estepozocargado[0])
		del(nortepozocargado[0])
		del(cotaterrenocargado[0])
		del(cotafondocargado[0])
		del(pozoinicialcargado[0])
		del(pozofinalcargado[0])
		del(diametrocargado[0])
		del(rugosidadcargado[0])
		del(claveinicialcargado[0])
		del(clavefinalcargado[0])
		del(IDdescargacargado[0])
		del(estedescargacargado[0])
		del(nortedescargacargado[0])
		del(cotafondodescargacargado[0])

	#crear listas de nodos y coordenadas completas (pozos+descargas)

		lista_nodos_total=IDPozocargado+IDdescargacargado
		lista_este_total=estepozocargado+estedescargacargado
		lista_norte_total=nortepozocargado+nortedescargacargado

		#print(lista_nodos_total)
		#print(lista_este_total)
		#print(lista_norte_total)

	#calcular de elementos en la lista
		
		global longitud_lista_pozos
		longitud_lista_pozos = len(IDPozocargado)
		cantidad_pozos=range(longitud_lista_pozos)
		#print(longitud_lista_pozos)

		global longitud_lista_tramos
		longitud_lista_tramos = len(pozoinicialcargado)
		cantidad_pozos=range(longitud_lista_tramos)
		#print(longitud_lista_tramos)

		global longitud_lista_descarga
		longitud_lista_descarga = len(IDdescargacargado)
		cantidad_pozos=range(longitud_lista_descarga)
		#print(longitud_lista_descarga)


		#Crear lista de caudales
		if (tipo_caudal_seleccionado==1):
			for row in listahojatramos.iter_rows(min_col=columnacaudal, max_col=columnacaudal):
				caudalcargado.extend([celda.value for celda in row])

		else:

			calcular_caudal_puntual()

		if (tipo_caudal_seleccionado==1):
			del(caudalcargado[0])

	#Crear el archivo inp
		file=open("Modelo SWMM-EXCEL.inp","w")
		file.write("[TITLE]"+"\n")
		file.write(";;Project Title/Notes"+"\n")

		file.write("\n")

		file.write("[OPTIONS]"+"\n")
		file.write(";;Option             Value"+"\n")
		file.write("FLOW_UNITS           LPS"+"\n")
		file.write("INFILTRATION         HORTON"+"\n")
		file.write("FLOW_ROUTING         STEADY"+"\n")
		file.write("LINK_OFFSETS         ELEVATION"+"\n")
		file.write("MIN_SLOPE            0"+"\n")
		file.write("ALLOW_PONDING        YES"+"\n")
		file.write("SKIP_STEADY_STATE    NO"+"\n")

		file.write("\n")

		file.write("START_DATE           01/09/2021"+"\n")
		file.write("START_TIME           00:00:00"+"\n")
		file.write("REPORT_START_DATE    01/09/2021"+"\n")
		file.write("REPORT_START_TIME    00:00:00"+"\n")
		file.write("END_DATE             01/09/2021"+"\n")
		file.write("END_TIME             06:00:00"+"\n")
		file.write("SWEEP_START          1/1"+"\n")
		file.write("SWEEP_END            12/31"+"\n")
		file.write("DRY_DAYS             0"+"\n")
		file.write("REPORT_STEP          00:15:00"+"\n")
		file.write("WET_STEP             00:05:00"+"\n")
		file.write("DRY_STEP             01:00:00"+"\n")
		file.write("ROUTING_STEP         0:00:30"+"\n")

		file.write("\n")

		file.write("INERTIAL_DAMPING     PARTIAL"+"\n")
		file.write("NORMAL_FLOW_LIMITED  BOTH"+"\n")
		file.write("FORCE_MAIN_EQUATION  H-W"+"\n")
		file.write("VARIABLE_STEP        0.75"+"\n")
		file.write("LENGTHENING_STEP     0"+"\n")
		file.write("MIN_SURFAREA         0"+"\n")
		file.write("MAX_TRIALS           0"+"\n")
		file.write("HEAD_TOLERANCE       0"+"\n")
		file.write("SYS_FLOW_TOL         5"+"\n")
		file.write("LAT_FLOW_TOL         5"+"\n")
		file.write("MINIMUM_STEP         0.5"+"\n")
		file.write("THREADS              1"+"\n")
		
		file.write("\n")

		file.write("[EVAPORATION]"+"\n")
		file.write(";;Data Source    Parameters"+"\n")
		file.write(";;-------------- ----------------"+"\n")
		file.write("CONSTANT         0.0"+"\n")
		file.write("DRY_ONLY         NO"+"\n")

		file.write("\n")

		file.write("[JUNCTIONS]"+"\n")
		file.write(";;Name           Elevation  MaxDepth   InitDepth  SurDepth   Aponded   "+"\n")
		file.write(";;-------------- ---------- ---------- ---------- ---------- ----------"+"\n")

		for i in range(longitud_lista_pozos):
			file.write(str(IDPozocargado[i])+"                "+str(cotafondocargado[i])+"          "+str(round(cotaterrenocargado[i]-cotafondocargado[i],3))+"          0          0          0          0         "+"\n")
		
		file.write("\n")

		file.write("[OUTFALLS]"+"\n")
		file.write(";;Name           Elevation  Type       Stage Data       Gated    Route To"+"\n")
		file.write(";;-------------- ---------- ---------- ---------------- -------- ----------------"+"\n")
		for i in range(longitud_lista_descarga):
			file.write(str(IDdescargacargado[i])+"                "+str(cotafondodescargacargado[i])+"          "+"FREE"+"                        NO                      "+"\n")

		file.write("\n")

		file.write("[CONDUITS]"+"\n")
		file.write(";;Name           From Node        To Node          Length     Roughness  InOffset   OutOffset  InitFlow   MaxFlow"+"\n")
		file.write(";;-------------- ---------------- ---------------- ---------- ---------- ---------- ---------- ---------- ----------"+"\n")
		for i in range(longitud_lista_tramos):
			
			try:
				longituddeltramo = str(round((((lista_norte_total[lista_nodos_total.index(pozofinalcargado[i])]-lista_norte_total[lista_nodos_total.index(pozoinicialcargado[i])])**2)+((lista_este_total[lista_nodos_total.index(pozofinalcargado[i])]-lista_este_total[lista_nodos_total.index(pozoinicialcargado[i])])**2))**(0.5),3))
			except ValueError:
				longituddeltramo = str(100)
				messagebox.showwarning(message="Falta incluir la informacion de alguno de los siguientes nodos en el excel: "+pozoinicialcargado[i]+"-"+pozofinalcargado[i]+" .Por lo anterior no se creo este tramo", title="Alerta!")
			
			file.write((str(i+1))+"                "+str(pozoinicialcargado[i])+"                "+str(pozofinalcargado[i])+"                "+longituddeltramo+"        "+str(rugosidadcargado[i])+"       "+str(round(claveinicialcargado[i]-diametrocargado[i],3))+"          "+str(round(clavefinalcargado[i]-diametrocargado[i],3))+"          0          0        "+"\n")
			
			


		file.write("\n")

		file.write("[XSECTIONS]"+"\n")
		file.write(";;Link           Shape        Geom1            Geom2      Geom3      Geom4      Barrels    Culvert"+"\n")
		file.write(";;-------------- ------------ ---------------- ---------- ---------- ---------- ---------- ----------"+"\n")
		for i in range(longitud_lista_tramos):
			file.write((str(i+1))+"                CIRCULAR     "+str(round(diametrocargado[i],3))+"              0          0          0          1                    "+"\n")


		file.write("\n")

		if (tipo_caudal_seleccionado==1):
			file.write("[INFLOWS]"+"\n")
			file.write(";;Node           Constituent      Time Series      Type     Mfactor  Sfactor  Baseline Pattern"+"\n")
			file.write(";;-------------- ---------------- ---------------- -------- -------- -------- -------- --------"+"\n")
			for i in range(longitud_lista_tramos):
				file.write(str(pozoinicialcargado[i])+'		FLOW		""		FLOW		1.0'+"		"+"1.0"+"		"+str(round(caudalcargado[i],3))+"		"+"\n")

		else:
			file.write("[INFLOWS]"+"\n")
			file.write(";;Node           Constituent      Time Series      Type     Mfactor  Sfactor  Baseline Pattern"+"\n")
			file.write(";;-------------- ---------------- ---------------- -------- -------- -------- -------- --------"+"\n")
			for i in range(longitud_lista_tramos):
				file.write(str(pozoinicialcargado[i])+'		FLOW		""		FLOW		1.0'+"		"+"1.0"+"		"+str(round(caudalcargadopuntual[i],3))+"		"+"\n")


		file.write("\n")

		file.write("[REPORT]"+"\n")
		file.write(";;Reporting Options"+"\n")
		file.write("INPUT      NO"+"\n")
		file.write("CONTROLS   NO"+"\n")
		file.write("SUBCATCHMENTS ALL"+"\n")
		file.write("NODES ALL"+"\n")
		file.write("LINKS ALL"+"\n")

		file.write("\n")

		file.write("[TAGS]"+"\n")

		file.write("\n")

		#calcular las dimensiones del mapa

		xmenor=estepozocargado[0]
		xmayor=estepozocargado[0]
		ymenor=nortepozocargado[0]
		ymayor=nortepozocargado[0]

		for i in range(longitud_lista_pozos):
			
			if (estepozocargado[i]<xmenor):
				xmenor=estepozocargado[i]

			if (estepozocargado[i]>xmayor):
				xmayor=estepozocargado[i]

			if (nortepozocargado[i]<ymenor):
				ymenor=nortepozocargado[i]

			if (nortepozocargado[i]>ymayor):
				ymayor=nortepozocargado[i]


		file.write("[MAP]"+"\n")
		file.write("DIMENSIONS "+str(round(xmenor,2)-25)+" "+str(round(ymenor,2)-173)+" "+str(round(xmayor,2)+25)+" "+str(round(ymayor,2)+173)+"\n")
		file.write("Units      Meters"+"\n")

		file.write("\n")

		file.write("[COORDINATES]"+"\n")
		file.write(";;Node           X-Coord            Y-Coord           "+"\n")
		file.write(";;-------------- ------------------ ------------------"+"\n")


		for i in range(longitud_lista_pozos):
			file.write(str(IDPozocargado[i])+"                "+str(estepozocargado[i])+"          "+str(nortepozocargado[i])+"          "+"\n")
		
		for i in range(longitud_lista_descarga):
			file.write(str(IDdescargacargado[i])+"                "+str(estedescargacargado[i])+"          "+str(nortedescargacargado[i])+"          "+"\n")	

		file.write("\n")

		file.write("[VERTICES]"+"\n")
		file.write(";;Link           X-Coord            Y-Coord           "+"\n")
		file.write(";;-------------- ------------------ ------------------"+"\n")

		file.close()

		messagebox.showinfo(parent=ventmodel,message="Modelo creado con éxito en la misma ruta del programa", title="Resultado")

def dibujar2d():
	pass

def dibujar3d():
	pass

#crea la ventana de trabajo	

ventana = tkinter.Tk()
ventana.geometry("800x600+450+200")
ventana.title("SWMM-EXCEL Alpha")
ventana.iconbitmap(r"C:\Users\Ing Juan Manuel\Documents\Sublime Text Ejercicios\SWMM-EXCEL\Icono.ico")
ventana.resizable(0,0)
imagen_portada = PhotoImage(file=r"C:\Users\Ing Juan Manuel\Documents\Sublime Text Ejercicios\SWMM-EXCEL\Portada.gif")
fondo=Label(ventana,image=imagen_portada).place(x=-2,y=-3)

#Crear Barra Menu

barraMenu=Menu(ventana)

#Crear los menus

mnucrearmodelo = Menu(barraMenu, tearoff = 0)
mnudibujar = Menu(barraMenu, tearoff = 0)
mnuayuda=Menu(barraMenu, tearoff = 0)

#Crear los comandos de los menus

mnucrearmodelo.add_command(label="Crear modelo SWMM-EXCEL",command=ventana_modelo)
mnucrearmodelo.add_separator()
mnucrearmodelo.add_command(label="Salir",command=ventana.quit)

######

mnudibujar.add_command(label="Dibujar Alcantarillado 2D",command=dibujar2d)
mnudibujar.add_command(label="Dibujar Alcantarillado 3D",command=dibujar3d)

#Agregar los menus a la barra de menus

barraMenu.add_cascade(label="Crear Modelo",menu=mnucrearmodelo)
barraMenu.add_cascade(label="Dibujar en AutoCAD",menu=mnudibujar)
barraMenu.add_cascade(label="Ayuda",menu=mnuayuda)

#Indicamos que la barra de menus estará en la ventana

ventana.config(menu=barraMenu)

#creador etiquetas

#autor = tkinter.Label(ventana, text= "Ing. Juan Manuel Pérez - Bogotá Colombia - Youtube: Juan Manuel Perez Hidráulica").place(x=200,y=550)


ventana.mainloop()



